#!/usr/bin/env python3
"""
Single-keyword LinkedIn job scraper with Excel export.

Uses li_at session cookie for authenticated access to LinkedIn search HTML
and the Voyager API for per-job detail enrichment.
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SEARCH_DELAY = 1.5   # seconds between search page fetches
DETAIL_DELAY = 0.8   # seconds between Voyager API calls

BASE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
}

VOYAGER_HEADERS = {
    "User-Agent": BASE_HEADERS["User-Agent"],
    "Accept": "application/vnd.linkedin.normalized+json+2.1",
    "Accept-Language": "en-US,en;q=0.9",
    "x-li-lang": "en_US",
    "x-restli-protocol-version": "2.0.0",
}


def load_cookie() -> str:
    load_dotenv()
    li_at = os.getenv("li_at")
    if not li_at:
        sys.exit("❌ li_at not found in .env — add it and retry.")
    return li_at


def make_session(li_at: str) -> requests.Session:
    session = requests.Session()
    session.headers.update(BASE_HEADERS)
    session.cookies.set("li_at", li_at, domain=".linkedin.com")
    try:
        session.get("https://www.linkedin.com/", timeout=15)
    except Exception:
        pass
    return session


def get_csrf_token(session: requests.Session) -> str:
    """CSRF token for Voyager API is the JSESSIONID cookie value (without quotes)."""
    return session.cookies.get("JSESSIONID", "").strip('"')


def search_page(session: requests.Session, keywords: str, geo_id: str,
                distance: int, time_filter: str, start: int) -> list[dict]:
    """Fetch one page of search results and return job stubs."""
    params: dict = {
        "keywords": keywords,
        "f_TPR": time_filter,
        "start": start,
    }
    if geo_id:
        params["geoId"] = geo_id
        params["distance"] = distance

    try:
        resp = session.get(
            "https://www.linkedin.com/jobs/search/",
            params=params,
            timeout=20,
        )
    except Exception as e:
        print(f"  ⚠️  Request error at start={start}: {e}")
        return []

    if resp.status_code == 429:
        print("  ⚠️  Rate limited (429). Sleeping 30s...")
        time.sleep(30)
        return []
    if resp.status_code != 200:
        print(f"  ⚠️  HTTP {resp.status_code} at start={start}")
        return []

    return _parse_jobs_from_html(resp.text)


def _parse_jobs_from_html(html: str) -> list[dict]:
    """Extract job stubs from the JSON embedded in LinkedIn's <code> blocks."""
    jobs = []
    for block in re.findall(r"<code[^>]*>(.*?)</code>", html, re.DOTALL):
        try:
            text = block.replace("&quot;", '"').replace("&amp;", "&").replace("&#39;", "'")
            data = json.loads(text)
        except (json.JSONDecodeError, ValueError):
            continue
        jobs.extend(_extract_job_stubs(data))
    return jobs


def _extract_job_stubs(data: dict) -> list[dict]:
    """Pull job ID, title, company, and location out of a LinkedIn JSON blob."""
    results = []
    if not isinstance(data, dict):
        return results

    for item in data.get("included", []):
        if not isinstance(item, dict):
            continue
        urn = item.get("entityUrn", "")
        if "jobPosting:" not in urn:
            continue

        job_id = urn.split("jobPosting:")[-1].split(",")[0].strip()
        title = item.get("title", "")
        if not job_id or not title:
            continue

        company = ""
        comp_details = item.get("companyDetails", {})
        if isinstance(comp_details, dict):
            for val in comp_details.values():
                if isinstance(val, dict):
                    res = val.get("companyResolutionResult", {})
                    if isinstance(res, dict):
                        company = res.get("name", "")
                    if not company:
                        company = val.get("name", "")
                if company:
                    break

        results.append({
            "job_id": job_id,
            "title": title,
            "company": company,
            "location": item.get("formattedLocation", ""),
        })

    return results


def get_job_details(session: requests.Session, job_id: str, csrf: str) -> dict:
    """Call Voyager API for per-job rich fields."""
    url = f"https://www.linkedin.com/voyager/api/jobs/jobPostings/{job_id}"
    headers = {
        **VOYAGER_HEADERS,
        "Csrf-Token": csrf,
        "Cookie": (
            f'li_at={session.cookies.get("li_at")}; '
            f'JSESSIONID="{csrf}"'
        ),
    }
    try:
        resp = session.get(
            url,
            params={"decorationId": "com.linkedin.voyager.deco.jobs.web.shared.WebFullJobPosting-65"},
            headers=headers,
            timeout=15,
        )
    except Exception as e:
        print(f"    ⚠️  Detail request failed for {job_id}: {e}")
        return {}

    if resp.status_code == 429:
        print("    ⚠️  Rate limited on detail API. Sleeping 30s...")
        time.sleep(30)
        return {}
    if resp.status_code != 200:
        return {}

    try:
        return _parse_voyager(resp.json(), job_id)
    except Exception:
        return {}


def _extract_company(main: dict, included: list, urn_map: dict) -> str:
    comp = main.get("companyDetails", {})
    if isinstance(comp, dict):
        for key, val in comp.items():
            if not isinstance(val, dict):
                continue
            # Unaffiliated jobs store the name directly
            if val.get("companyName"):
                return val["companyName"]
            # Affiliated jobs nest it under companyResolutionResult
            res = val.get("companyResolutionResult", {})
            if isinstance(res, dict) and res.get("name"):
                return res["name"]
            # Sometimes the resolution object is in included, keyed by company URN
            company_urn = val.get("company", "")
            if company_urn and company_urn in urn_map:
                name = urn_map[company_urn].get("name", "")
                if name:
                    return name

    # Last resort: scan included for any mini-company object
    for item in included:
        if not isinstance(item, dict):
            continue
        urn = item.get("entityUrn", "")
        if ("company" in urn.lower() or "miniCompany" in urn) and item.get("name"):
            return item["name"]

    return ""


def _extract_salary(main: dict) -> str:
    def _fmt(low, high, currency, period):
        low_s = str(low) if low else ""
        high_s = str(high) if high else ""
        if not low_s and not high_s:
            return ""
        parts = [currency, f"{low_s}–{high_s}"]
        if period:
            parts.append(period)
        return " ".join(parts).strip()

    # salaryInsights.compensationRange (most common)
    insight = main.get("salaryInsights", {})
    if isinstance(insight, dict):
        cr = insight.get("compensationRange", {})
        if isinstance(cr, dict):
            low = cr.get("minSalary", {})
            high = cr.get("maxSalary", {})
            currency = cr.get("currencyCode", "USD")
            period = cr.get("payPeriod", "")
            low_val = low.get("amount") if isinstance(low, dict) else low
            high_val = high.get("amount") if isinstance(high, dict) else high
            s = _fmt(low_val, high_val, currency, period)
            if s:
                return s

    # salary.compensationRange (alternate shape)
    salary_field = main.get("salary", {})
    if isinstance(salary_field, dict):
        cr = salary_field.get("compensationRange", {})
        if isinstance(cr, dict):
            low_val = cr.get("min") or cr.get("minimum") or cr.get("minSalary")
            high_val = cr.get("max") or cr.get("maximum") or cr.get("maxSalary")
            currency = cr.get("currencyCode", "USD")
            period = cr.get("payPeriod", "")
            s = _fmt(low_val, high_val, currency, period)
            if s:
                return s

    # compensationInfo
    comp_info = main.get("compensationInfo", {})
    if isinstance(comp_info, dict):
        low_val = comp_info.get("min") or comp_info.get("minimum")
        high_val = comp_info.get("max") or comp_info.get("maximum")
        currency = comp_info.get("currencyCode", "USD")
        period = comp_info.get("payPeriod", "")
        s = _fmt(low_val, high_val, currency, period)
        if s:
            return s

    return ""


def _parse_voyager(data: dict, job_id: str) -> dict:
    """Map Voyager API response fields to our flat job dict."""
    main = data.get("data", {})
    included = data.get("included", [])

    urn_map = {
        item.get("entityUrn", ""): item
        for item in included
        if isinstance(item, dict)
    }

    result: dict = {
        "job_id": job_id,
        "url": f"https://www.linkedin.com/jobs/view/{job_id}/",
        "title": main.get("title", ""),
        "location": main.get("formattedLocation", ""),
    }

    # Company — LinkedIn uses two different companyDetails shapes:
    #   jobPostingCompany      → companyResolutionResult.name  (affiliated)
    #   jobPostingUnaffiliated → companyName                   (unaffiliated)
    result["company"] = _extract_company(main, included, urn_map)

    # Work type (Remote / Hybrid / On-site)
    wpt = main.get("workplaceTypesResolutionResults", {})
    if isinstance(wpt, dict):
        result["work_type"] = ", ".join(
            v.get("localizedName", "")
            for v in wpt.values()
            if isinstance(v, dict) and v.get("localizedName")
        )
    else:
        result["work_type"] = ""

    # Employment type — try via URN elements list
    emp_type = ""
    for urn in main.get("employmentTypes", {}).get("*elements", []):
        et = urn_map.get(urn, {})
        if isinstance(et, dict) and et.get("localizedName"):
            emp_type = et["localizedName"]
            break
    # Fallback: scan included for known employment type names
    if not emp_type:
        known = {"Full-time", "Part-time", "Contract", "Temporary", "Internship", "Volunteer"}
        for item in included:
            if isinstance(item, dict) and item.get("localizedName") in known:
                emp_type = item["localizedName"]
                break
    result["employment_type"] = emp_type

    # Seniority
    seniority = ""
    for item in included:
        if isinstance(item, dict) and "seniorityLevel" in item.get("entityUrn", ""):
            seniority = item.get("localizedName", "")
            break
    result["seniority"] = seniority

    # Date posted (listedAt is epoch milliseconds)
    listed_at = main.get("listedAt", 0)
    if listed_at:
        try:
            result["date_posted"] = datetime.utcfromtimestamp(listed_at / 1000).strftime("%Y-%m-%d")
        except Exception:
            result["date_posted"] = ""
    else:
        result["date_posted"] = ""

    result["salary"] = _extract_salary(main)

    # Easy apply
    apply = main.get("applyMethod", {})
    result["easy_apply"] = (
        "Yes"
        if (
            isinstance(apply, dict)
            and "com.linkedin.voyager.jobs.ComplexOnsiteApply" in apply
        )
        or main.get("easyApplyUrl")
        else "No"
    )

    # Applicants
    applies = main.get("applies", None)
    result["applicants"] = applies if applies else ""

    # Description (first 2000 chars)
    desc = main.get("description", {})
    if isinstance(desc, dict):
        desc_text = desc.get("text", "")
    else:
        desc_text = str(desc) if desc else ""
    result["description"] = desc_text[:2000]

    return result


def export_to_excel(jobs: list[dict], output_path: str) -> None:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    columns = [
        ("Job ID", 15),
        ("Job Title", 40),
        ("Company", 30),
        ("Location", 25),
        ("Work Type", 15),
        ("Employment Type", 18),
        ("Seniority", 18),
        ("Date Posted", 14),
        ("Salary", 25),
        ("Easy Apply", 12),
        ("Applicants", 12),
        ("URL", 50),
        ("Description", 60),
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    url_font = Font(color="0563C1", underline="single")

    for row_idx, job in enumerate(jobs, start=2):
        url = job.get("url") or f"https://www.linkedin.com/jobs/view/{job.get('job_id', '')}/"
        row_values = [
            job.get("job_id", ""),
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("work_type", ""),
            job.get("employment_type", ""),
            job.get("seniority", ""),
            job.get("date_posted", ""),
            job.get("salary", ""),
            job.get("easy_apply", ""),
            job.get("applicants", ""),
            url,
            job.get("description", ""),
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 12 and url:
                cell.hyperlink = url
                cell.font = url_font
            cell.alignment = Alignment(
                wrap_text=(col_idx == 13),
                vertical="top",
            )

    wb.save(output_path)
    print(f"✅ Saved {len(jobs)} jobs → {output_path}")


def scrape(
    keywords: str,
    geo_id: str,
    distance: int,
    time_filter: str,
    max_jobs: int,
    fetch_details: bool,
    output: str | None,
) -> list[dict]:
    """Full scraping pipeline. Returns list of job dicts."""
    li_at = load_cookie()
    session = make_session(li_at)
    csrf = get_csrf_token(session)

    if not csrf:
        print("⚠️  Could not get CSRF token — detail API calls may fail.")

    print(f"🔍 '{keywords}' | geo={geo_id or 'any'} | dist={distance}mi | filter={time_filter}")

    seen_ids: set[str] = set()
    all_jobs: list[dict] = []
    start = 0
    page_size = 14
    consecutive_empty = 0

    while len(all_jobs) < max_jobs:
        print(f"  Page start={start:>4} | collected={len(all_jobs)}", end="\r")
        page_jobs = search_page(session, keywords, geo_id, distance, time_filter, start)
        new_jobs = [j for j in page_jobs if j.get("job_id") not in seen_ids]

        if not new_jobs:
            consecutive_empty += 1
            if consecutive_empty >= 3:
                print(f"\n  ℹ️  No new results after {consecutive_empty} empty pages. Stopping.")
                break
        else:
            consecutive_empty = 0

        for job in new_jobs:
            if len(all_jobs) >= max_jobs:
                break
            seen_ids.add(job["job_id"])
            all_jobs.append(job)

        start += page_size
        time.sleep(SEARCH_DELAY)

    print(f"\n✅ Found {len(all_jobs)} unique jobs")

    if fetch_details and all_jobs:
        print(f"📋 Fetching details for {len(all_jobs)} jobs...")
        for i, job in enumerate(all_jobs):
            details = get_job_details(session, job["job_id"], csrf)
            if details:
                job.update(details)
            if (i + 1) % 10 == 0 or i + 1 == len(all_jobs):
                print(f"  Detail {i+1}/{len(all_jobs)}", end="\r")
            time.sleep(DETAIL_DELAY)
        print()

    for job in all_jobs:
        job.setdefault("url", f"https://www.linkedin.com/jobs/view/{job['job_id']}/")

    if output:
        export_to_excel(all_jobs, output)

    return all_jobs


def main() -> None:
    parser = argparse.ArgumentParser(description="Scrape LinkedIn job listings")
    parser.add_argument("--keywords", default="infrastructure")
    parser.add_argument("--geo-id", default="101098412", help="LinkedIn geo ID (default: Massachusetts)")
    parser.add_argument("--distance", type=int, default=25)
    parser.add_argument("--time-filter", default="r604800",
                        help="r86400=24h | r604800=1 week | r2592000=1 month")
    parser.add_argument("--max-jobs", type=int, default=650)
    parser.add_argument("--no-details", action="store_true", help="Skip per-job Voyager API calls")
    parser.add_argument("--output", help="Output .xlsx path (default: auto-named in .tmp/)")
    args = parser.parse_args()

    if not args.output:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        kw_slug = args.keywords.replace(" ", "_")
        args.output = f".tmp/linkedin_jobs_{kw_slug}_{ts}.xlsx"

    scrape(
        keywords=args.keywords,
        geo_id=args.geo_id,
        distance=args.distance,
        time_filter=args.time_filter,
        max_jobs=args.max_jobs,
        fetch_details=not args.no_details,
        output=args.output,
    )


if __name__ == "__main__":
    main()
