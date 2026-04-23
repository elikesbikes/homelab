#!/usr/bin/env python3
"""
TrainerRoad workout analyzer.
Reads the latest .tmp/trainerroad_activities_*.xlsx and outputs structured
JSON to stdout for the /trainerroad-insights skill to interpret.

Usage:
    python3 tools/analyze_workouts.py
    python3 tools/analyze_workouts.py --weeks 8
    python3 tools/analyze_workouts.py --file .tmp/trainerroad_activities_20240421_1200.xlsx
"""

import argparse
import glob
import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl


# ── Load data ────────────────────────────────────────────────────────────────

def find_latest_xlsx() -> str:
    files = glob.glob(".tmp/trainerroad_activities_*.xlsx")
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def load_activities(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    activities = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        act = {}
        for header, value in zip(headers, row):
            if header and value is not None:
                key = str(header).lower().replace(" ", "_")
                act[key] = value
        if act:
            activities.append(act)
    return activities


# ── Analysis helpers ─────────────────────────────────────────────────────────

def safe_float(val):
    try:
        return float(val) if val is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def week_start(date_str: str) -> str:
    try:
        d = datetime.strptime(date_str[:10], "%Y-%m-%d")
        monday = d - timedelta(days=d.weekday())
        return monday.strftime("%Y-%m-%d")
    except Exception:
        return "unknown"


def filter_by_weeks(activities: list[dict], weeks: int) -> list[dict]:
    cutoff = datetime.now() - timedelta(weeks=weeks)
    result = []
    for act in activities:
        try:
            d = datetime.strptime(act.get("date", "")[:10], "%Y-%m-%d")
            if d >= cutoff:
                result.append(act)
        except Exception:
            pass
    return result


def weekly_tss(activities: list[dict]) -> list[dict]:
    weekly = defaultdict(float)
    for act in activities:
        w = week_start(act.get("date", ""))
        weekly[w] += safe_float(act.get("tss"))
    return [{"week": k, "tss": round(v, 1)} for k, v in sorted(weekly.items())]


def workout_type_breakdown(activities: list[dict]) -> dict:
    counts = defaultdict(int)
    tss_by_type = defaultdict(float)
    for act in activities:
        wtype = str(act.get("workout_type") or "Unknown").strip()
        counts[wtype] += 1
        tss_by_type[wtype] += safe_float(act.get("tss"))
    return {
        t: {"count": counts[t], "total_tss": round(tss_by_type[t], 1)}
        for t in counts
    }


def top_workouts(activities: list[dict], n: int = 5) -> list[dict]:
    sorted_acts = sorted(activities, key=lambda a: safe_float(a.get("tss")), reverse=True)
    return [
        {
            "date": a.get("date", ""),
            "name": a.get("name", ""),
            "tss": safe_float(a.get("tss")),
            "duration_min": safe_float(a.get("duration_min")),
        }
        for a in sorted_acts[:n]
    ]


def consistency_streak(activities: list[dict]) -> dict:
    if not activities:
        return {"current_streak_days": 0, "longest_streak_days": 0}

    dates = set()
    for act in activities:
        try:
            d = datetime.strptime(act.get("date", "")[:10], "%Y-%m-%d").date()
            dates.add(d)
        except Exception:
            pass

    today = datetime.now().date()
    sorted_dates = sorted(dates, reverse=True)

    # Current streak (counting from today or yesterday)
    current = 0
    check = today
    for _ in range(365):
        if check in dates:
            current += 1
            check -= timedelta(days=1)
        elif check == today:
            check -= timedelta(days=1)
        else:
            break

    # Longest streak
    longest = 0
    streak = 0
    prev = None
    for d in sorted(dates):
        if prev is None or (d - prev).days == 1:
            streak += 1
        else:
            longest = max(longest, streak)
            streak = 1
        prev = d
    longest = max(longest, streak)

    return {"current_streak_days": current, "longest_streak_days": longest}


def tss_ramp_rate(weekly_data: list[dict]) -> float:
    """Week-over-week TSS change over the last 4 weeks."""
    if len(weekly_data) < 2:
        return 0.0
    recent = weekly_data[-4:]
    if len(recent) < 2:
        return 0.0
    first_tss = recent[0]["tss"]
    last_tss = recent[-1]["tss"]
    if first_tss == 0:
        return 0.0
    return round((last_tss - first_tss) / first_tss * 100, 1)


def power_trends(activities: list[dict]) -> dict:
    powered = [a for a in activities if safe_float(a.get("avg_power")) > 0]
    if not powered:
        return {}
    avg_powers = [safe_float(a.get("avg_power")) for a in powered]
    nps = [safe_float(a.get("np")) for a in powered if safe_float(a.get("np")) > 0]
    return {
        "avg_power_mean": round(sum(avg_powers) / len(avg_powers), 1),
        "avg_power_max": max(avg_powers),
        "np_mean": round(sum(nps) / len(nps), 1) if nps else None,
    }


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Analyze TrainerRoad workout data")
    parser.add_argument("--weeks", type=int, default=0, help="Limit analysis to last N weeks (0 = all)")
    parser.add_argument("--file",  default="",          help="Path to xlsx file (default: latest in .tmp/)")
    args = parser.parse_args()

    xlsx_path = args.file or find_latest_xlsx()
    if not xlsx_path:
        print(json.dumps({"error": "No xlsx file found in .tmp/. Run scrape_trainerroad_activities.py first."}))
        sys.exit(1)

    activities = load_activities(xlsx_path)
    if not activities:
        print(json.dumps({"error": "No activities found in file."}))
        sys.exit(1)

    analysis_activities = filter_by_weeks(activities, args.weeks) if args.weeks else activities
    weekly = weekly_tss(analysis_activities)
    breakdown = workout_type_breakdown(analysis_activities)
    power = power_trends(analysis_activities)
    ramp = tss_ramp_rate(weekly)
    streak = consistency_streak(analysis_activities)
    tops = top_workouts(analysis_activities, n=5)

    total_tss = sum(w["tss"] for w in weekly)
    avg_weekly_tss = round(total_tss / len(weekly), 1) if weekly else 0
    total_hours = round(sum(safe_float(a.get("duration_min")) for a in analysis_activities) / 60, 1)

    output = {
        "meta": {
            "file": xlsx_path,
            "total_activities": len(analysis_activities),
            "date_range": {
                "earliest": min(a["date"] for a in analysis_activities if a.get("date")),
                "latest": max(a["date"] for a in analysis_activities if a.get("date")),
            },
            "analysis_window_weeks": args.weeks or "all",
        },
        "summary": {
            "total_tss": round(total_tss, 1),
            "avg_weekly_tss": avg_weekly_tss,
            "tss_ramp_rate_pct": ramp,
            "total_training_hours": total_hours,
            "current_streak_days": streak["current_streak_days"],
            "longest_streak_days": streak["longest_streak_days"],
        },
        "weekly_tss": weekly[-12:],  # last 12 weeks max for readability
        "workout_type_breakdown": breakdown,
        "top_workouts_by_tss": tops,
        "power_trends": power,
    }

    print(json.dumps(output, indent=2))


if __name__ == "__main__":
    main()
