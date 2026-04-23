#!/usr/bin/env python3
"""
TrainerRoad Activity Scraper
Logs into TrainerRoad with Playwright, then uses Playwright's request context
(which inherits httpOnly auth cookies) to paginate the activities API.

Usage:
    python3 tools/scrape_trainerroad_activities.py
    python3 tools/scrape_trainerroad_activities.py --username elikesbikes --output .tmp/out.xlsx
    python3 tools/scrape_trainerroad_activities.py --no-headless   # show browser for debug
"""

import argparse
import asyncio
import os
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

BASE_URL    = "https://www.trainerroad.com"
APP_API     = f"{BASE_URL}/app/api"
PAGE_SIZE   = 20
PAGE_DELAY  = 1.0   # seconds between paginated requests


# ── Excel export ─────────────────────────────────────────────────────────────

COLUMNS = [
    ("date",              "Date",              14),
    ("name",              "Name",              40),
    ("sport",             "Sport",             14),
    ("activity_type",     "Activity Type",     20),
    ("workout_category",  "Workout Category",  22),
    ("duration_min",      "Duration (min)",    14),
    ("tss",               "TSS",                8),
    ("intensity_factor",  "IF",                 8),
    ("np",                "NP (watts)",        12),
    ("kj",                "kJ",                 8),
    ("distance_mi",       "Distance (mi)",     14),
    ("progression_level", "Progression Lvl",  16),
    ("survey",            "Survey",            16),
    ("is_external",       "External",          10),
    ("device",            "Device",            22),
    ("url",               "URL",               55),
]


def export_to_excel(activities: list[dict], output_path: str) -> None:
    if not activities:
        print("⚠️  No activities to export.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TrainerRoad Activities"

    header_fill = PatternFill("solid", fgColor="DA291C")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, (_, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    for row_idx, act in enumerate(activities, start=2):
        for col_idx, (field, _, _) in enumerate(COLUMNS, start=1):
            value = act.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if field == "url" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Saved {len(activities)} activities → {output_path}")


# ── Data parsing ──────────────────────────────────────────────────────────────

def _meters_to_miles(m) -> float | None:
    try:
        return round(float(m) / 1609.344, 2)
    except (TypeError, ValueError):
        return None


def _seconds_to_min(s) -> float | None:
    try:
        return round(float(s) / 60, 1)
    except (TypeError, ValueError):
        return None


# Maps TR WorkoutType integers to human-readable training zone labels.
# Source: TrainerRoad API (field WorkoutType on activity card models).
WORKOUT_TYPE_MAP = {
    1:  "Endurance",
    2:  "Sweet Spot",
    3:  "Threshold",
    4:  "VO2 Max",
    5:  "Anaerobic",
    6:  "Sprint",
    7:  "Recovery",
    8:  "Tempo",
    9:  "Over-Under",
    10: "Race",
}


def parse_activity(raw: dict, username: str, category_cache: dict = None) -> dict:
    date_raw = raw.get("Started") or raw.get("WorkoutDate") or ""
    try:
        date = datetime.fromisoformat(date_raw.rstrip("Z")).strftime("%Y-%m-%d") if date_raw else ""
    except Exception:
        date = date_raw

    activity_id = raw.get("Id") or raw.get("WorkoutRecordId") or ""
    raw_type = raw.get("$type", "")
    # Strip "CardModel" suffix for a cleaner label
    activity_type = raw_type.replace("ActivityCardModel", "").replace("CardModel", "")

    device_list = raw.get("Devices") or []
    device = device_list[0].get("Name", "") if device_list else ""

    tss = raw.get("Tss")
    if tss is not None:
        try:
            tss = round(float(tss), 1)
        except Exception:
            pass

    if_raw = raw.get("IntensityFactor")
    if_val = None
    if if_raw is not None:
        try:
            if_val = round(float(if_raw), 3)
        except Exception:
            pass

    # Resolve workout category: prefer live lookup cache, then API fields, then local map
    workout_category = ""
    workout_id = raw.get("WorkoutId")
    if category_cache and workout_id and workout_id in category_cache:
        workout_category = category_cache[workout_id]
    if not workout_category:
        wt_int = raw.get("WorkoutType")
        if wt_int is not None:
            try:
                workout_category = WORKOUT_TYPE_MAP.get(int(wt_int), f"Type {wt_int}")
            except (TypeError, ValueError):
                workout_category = str(wt_int)
    if not workout_category:
        workout_category = (
            raw.get("WorkoutTypeDescription")
            or raw.get("EnergySystem")
            or raw.get("TrainingType")
            or ""
        )

    return {
        "date":              date,
        "name":              raw.get("Name") or "",
        "sport":             raw.get("Sport") or "",
        "activity_type":     activity_type,
        "workout_category":  workout_category,
        "duration_min":      _seconds_to_min(raw.get("Duration")),
        "tss":               tss,
        "intensity_factor":  if_val,
        "np":                raw.get("NormalizedPower"),
        "kj":                raw.get("Kj"),
        "distance_mi":       _meters_to_miles(raw.get("Distance")),
        "progression_level": raw.get("ProgressionLevel"),
        "survey":            raw.get("SurveyOptionText") or raw.get("TranslatedSurveyOptionText") or "",
        "is_external":       raw.get("IsExternal"),
        "device":            device,
        "url":               f"{BASE_URL}/app/activities/{activity_id}" if activity_id else "",
    }


# ── Playwright scraper ────────────────────────────────────────────────────────

async def login(page, email: str, password: str) -> bool:
    print("🔐 Logging in to TrainerRoad...")
    await page.goto(f"{BASE_URL}/app/login", wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(2)

    try:
        await page.wait_for_selector("#username", timeout=10000)
        await page.fill("#username", email)
        await page.fill("#password", password)
        await page.locator('button[type="button"]').first.click()
        await asyncio.sleep(6)

        if "login" in page.url.lower():
            print("❌ Login failed — still on login page. Check TR_EMAIL/TR_PASSWORD in .env")
            return False

        print(f"   ✓ Logged in. Landing: {page.url}")
        return True
    except Exception as e:
        print(f"❌ Login error: {e}")
        return False


async def fetch_all_activities(context, username: str) -> list[dict]:
    """Paginate the activities API using Playwright's request context (httpOnly cookie aware)."""
    print(f"📡 Fetching activities for @{username}...")

    # Check totals first
    totals_resp = await context.request.get(f"{APP_API}/activities/{username}/totals")
    if totals_resp.status == 200:
        totals = await totals_resp.json()
        total = totals.get("TotalRides", "?")
        print(f"   Total activities on record: {total}")

    all_raw = []
    skip = 0

    while True:
        url = f"{APP_API}/activities/{username}/activities?skip={skip}&count={PAGE_SIZE}&activityFilter=0"
        resp = await context.request.get(url)

        if resp.status != 200:
            print(f"   ⚠️  API returned {resp.status} at skip={skip} — stopping.")
            break

        batch = await resp.json()
        if not batch:
            break

        all_raw.extend(batch)
        page_num = skip // PAGE_SIZE + 1
        print(f"   Page {page_num}: +{len(batch)} (total so far: {len(all_raw)})")

        if len(batch) < PAGE_SIZE:
            break

        skip += PAGE_SIZE
        await asyncio.sleep(PAGE_DELAY)

    return all_raw


async def fetch_workout_categories(context, workout_ids: set) -> dict:
    """Fetch workout type for each unique WorkoutId. Returns {workoutId: category_string}."""
    if not workout_ids:
        return {}
    cache = {}
    ids = sorted(workout_ids)
    print(f"🏋️  Fetching workout categories for {len(ids)} unique workouts...")
    for i, wid in enumerate(ids, 1):
        try:
            resp = await context.request.get(f"{APP_API}/workouts/{wid}")
            if resp.status == 200:
                data = await resp.json()
                category = (
                    data.get("WorkoutTypeName")
                    or data.get("EnergySystemName")
                    or data.get("WorkoutType", {}).get("Name", "") if isinstance(data.get("WorkoutType"), dict) else ""
                    or str(data.get("WorkoutTypeId", "")) if data.get("WorkoutTypeId") else ""
                )
                cache[wid] = category
            if i % 100 == 0:
                print(f"   {i}/{len(ids)} fetched...")
            await asyncio.sleep(0.3)
        except Exception:
            pass
    print(f"   ✓ Categories fetched for {len(cache)}/{len(ids)} workouts.")
    return cache


async def run(username: str, output_path: str, headless: bool = True,
              dump_raw: int = 0, dump_workout: int = 0) -> None:
    from playwright.async_api import async_playwright

    email    = os.getenv("TR_EMAIL")
    password = os.getenv("TR_PASSWORD")
    if not email or not password:
        raise SystemExit("❌ TR_EMAIL and TR_PASSWORD must be set in .env")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1440, "height": 900},
        )
        page = await context.new_page()

        if not await login(page, email, password):
            await browser.close()
            raise SystemExit("Aborting — login failed.")

        if dump_workout:
            import json
            resp = await context.request.get(f"{APP_API}/workouts/{dump_workout}")
            print(f"\n🔍 Raw workout API response for WorkoutId={dump_workout} (status {resp.status}):\n")
            if resp.status == 200:
                print(json.dumps(await resp.json(), indent=2))
            else:
                print(await resp.text())
            await browser.close()
            return

        raw_activities = await fetch_all_activities(context, username)

        if not raw_activities:
            await browser.close()
            raise SystemExit("❌ No activities returned by API.")

        if dump_raw:
            import json
            await browser.close()
            sample = raw_activities[:dump_raw]
            print(f"\n🔍 Raw API response for first {len(sample)} activit(y/ies):\n")
            print(json.dumps(sample, indent=2))
            return

        # Collect unique WorkoutIds from structured (non-external) TR workouts
        workout_ids = {
            raw.get("WorkoutId")
            for raw in raw_activities
            if raw.get("WorkoutId") and not raw.get("IsExternal")
        }
        category_cache = await fetch_workout_categories(context, workout_ids)
        await browser.close()

    print(f"\n📊 Parsing {len(raw_activities)} activities...")
    activities = [parse_activity(a, username, category_cache) for a in raw_activities]
    activities.sort(key=lambda a: a.get("date", ""), reverse=True)

    export_to_excel(activities, output_path)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Scrape TrainerRoad activities to Excel")
    parser.add_argument("--username",    default=os.getenv("TR_USERNAME", "elikesbikes"))
    parser.add_argument("--output",      default="")
    parser.add_argument("--no-headless", action="store_true", help="Show browser (debug)")
    parser.add_argument("--dump-raw",     type=int, default=0, metavar="N",
                        help="Print raw JSON for first N activities (skips Excel export) — use to inspect API fields")
    parser.add_argument("--dump-workout", type=int, default=0, metavar="WORKOUT_ID",
                        help="Print raw workout API JSON for a specific WorkoutId — use to discover zone field names")
    args = parser.parse_args()

    if not args.output:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        args.output = f".tmp/trainerroad_activities_{ts}.xlsx"

    asyncio.run(run(
        username=args.username,
        output_path=args.output,
        headless=not args.no_headless,
        dump_raw=args.dump_raw,
        dump_workout=args.dump_workout,
    ))


if __name__ == "__main__":
    main()
